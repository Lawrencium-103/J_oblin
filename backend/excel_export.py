from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

HEADER_FILL = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Calibri", size=10)
THIN_BORDER = Border(
    left=Side(style="thin", color="d1d5db"),
    right=Side(style="thin", color="d1d5db"),
    top=Side(style="thin", color="d1d5db"),
    bottom=Side(style="thin", color="d1d5db"),
)

COLUMNS = [
    ("ID", 6),
    ("Title", 45),
    ("Company", 25),
    ("Location", 20),
    ("Category", 18),
    ("Source", 16),
    ("Posted Date", 14),
    ("Date Found", 14),
    ("URL", 50),
    ("Description", 60),
    ("Match Score", 10),
    ("Is Graduate", 10),
    ("Has Full Info", 10),
]


def export_jobs_to_excel(jobs: list[dict]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "All Jobs"

    # Header row
    for col_idx, (col_name, width) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER
        ws.column_dimensions[chr(64 + col_idx)].width = width

    # Data rows
    for row_idx, job in enumerate(jobs, 2):
        values = [
            job.get("id", ""),
            job.get("title", ""),
            job.get("company", ""),
            job.get("location", ""),
            job.get("job_category", ""),
            job.get("source", ""),
            job.get("posted_date", ""),
            job.get("date_found", ""),
            job.get("url", ""),
            (job.get("description", "") or "")[:500],
            job.get("match_score", ""),
            "Yes" if job.get("is_graduate") else "No",
            "Yes" if job.get("has_full_info") else "No",
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = Alignment(vertical="top")

    # Freeze header row
    ws.freeze_panes = "A2"

    output_path = str(Path(__file__).resolve().parent.parent / "data" / "generated" / "jobs_export.xlsx")
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
